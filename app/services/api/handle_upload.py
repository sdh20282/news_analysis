from fastapi import UploadFile

from openpyxl import load_workbook
from io import BytesIO

import traceback

from app.services.upload.file import get_output_path, open_output_stream
from app.services.upload.validator import validate_row
from app.schemas.json_response import JSONResponse
from app.utils.response import success_response, error_response


async def validate_and_save_excel(file: UploadFile) -> JSONResponse[None]:
    try:
        if not file:
            raise Exception("파일이 전달되지 않았습니다.")

        contents = await file.read()
        workbook = load_workbook(filename=BytesIO(contents), read_only=True)
        sheet = workbook.active

        # 헤더 (1행) 검증
        header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        requiredColumns = {"번호", "본문", "긍정", "부정", "중립"}

        if not requiredColumns.issubset(set(header)):
            raise Exception(f"필수 컬럼 누락: {requiredColumns - set(header)}")

        colMap = {col: header.index(col) for col in requiredColumns}

        outputPath = get_output_path()
        valid_count = 0

        with open_output_stream(outputPath) as outFile:
            # 데이터 검증
            for rowIdx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                try:
                    jsonLine = validate_row(row, colMap)

                    if jsonLine is None:
                        continue

                    outFile.write(jsonLine + "\n")

                    valid_count += 1
                except Exception as e:
                    print(f"[SKIP] {rowIdx}행 오류: {e}")

                    continue

            print(f"유효 데이터 {valid_count}건")

            return success_response(
                message="엑셀 파싱 및 검증 성공",
            )
    except Exception as e:
        print("[ERROR] app/services/api/handle_upload.py 예외 발생:", e)
        print(f"File: {file.filename if file else 'No file'}")

        traceback.print_exc()

        return error_response(
            message="엑셀 파싱 및 검증 실패",
        )
